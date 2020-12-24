from openpyxl.chart import BarChart, BarChart3D, ScatterChart, Reference, Series


# makeChart module functions
# CAUTION: Please exit out of your excel sheet before executing any of these methods; otherwise, it won't work

def convert_to_Scatter(sheet_name, wb_obj, col_num_x=1, col_num_y=2, min_row_num=2, max_row_num=None,
                       x_axis_title='X-AXIS',
                       y_axis_title="Y-AXIS",
                       chart_title="chart-1"):
    sheet_1 = wb_obj[sheet_name]
    if max_row_num is None:
        max_row_num = sheet_1.max_row

    chart_1 = ScatterChart()
    chart_1.title = chart_title
    ref_obj_x = Reference(sheet_1, min_col=col_num_x, max_col=col_num_x, min_row=min_row_num, max_row=max_row_num)
    ref_obj_y = Reference(sheet_1, min_col=col_num_y, max_col=col_num_y, min_row=min_row_num, max_row=max_row_num)
    chart_1.x_axis.title = x_axis_title
    chart_1.y_axis.title = y_axis_title
    series = Series(ref_obj_y, ref_obj_x)
    chart_1.series.append(series)
    sheet_1.add_chart(chart_1, sheet_1.cell(2, sheet_1.max_column + 1).coordinate)


def convert_to_BarChart(sheet_name, wb_obj, min_col_num, max_col_num, min_row_num, max_row_num, min_cat_col,
                        min_cat_row, max_cat_row, chart_title="Chart 1", x_ax_title="X-Axis", y_ax_title="Y-Axis"):
    chart_1 = BarChart()

    sheet_1 = wb_obj[sheet_name]
    chart_1.title = chart_title
    chart_1.y_axis.title = x_ax_title
    chart_1.x_axis.title = y_ax_title
    data = Reference(sheet_1, min_col=min_col_num, min_row=min_row_num, max_row=max_row_num, max_col=max_col_num)
    cats = Reference(sheet_1, min_col=min_cat_col, min_row=min_cat_row, max_row=max_cat_row)
    chart_1.add_data(data, titles_from_data=True)
    chart_1.set_categories(cats)
    sheet_1.add_chart(chart_1, sheet_1.cell(2, sheet_1.max_column + 1).coordinate)


def convert_to_BarChart3D(sheet_name, wb_obj, chart_title="Chart 1"):
    chart_1 = BarChart3D()
    chart_1.title = chart_title
    sheet_1 = wb_obj[sheet_name]
    my_values = Reference(sheet_1, min_col=2, min_row=1, max_col=sheet_1.max_column, max_row=sheet_1.max_row)
    titles = Reference(sheet_1, min_col=1, min_row=2, max_row=sheet_1.max_row)
    chart_1.add_data(data=my_values, titles_from_data=True)
    chart_1.set_categories(titles)
    sheet_1.add_chart(chart_1, sheet_1.cell(2, sheet_1.max_column + 1).coordinate)


class DataSheet:
    """A Data Sheet object is simply a portion of a data table """

    def __init__(self, sheet_name, wb_obj):
        self.sheet_ID = sheet_name
        self.wb_obj = wb_obj
        try:
            self.sheet = wb_obj[sheet_name]
        except KeyError:
            self.sheet = wb_obj.create_sheet(sheet_name)
            print("New empty sheet created")

    def get_suitable_chart(self, start_row_first_cell, end_row_num, chart_title_name):

        # will automatically call convert_to_BarChart3D() if I have a DataSheet with a portion of data in Sheet12

        for i in self.sheet[start_row_first_cell:(start_row_first_cell[0] + str(end_row_num))]:

            if type(i[0].value) != str:
                return "pick a different chart "

        return convert_to_BarChart3D(self.sheet_ID, self.wb_obj, chart_title=chart_title_name)
